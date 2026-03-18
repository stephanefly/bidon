import os
import re
from datetime import datetime
from typing import Dict, List

import h5py
import numpy as np
import pandas as pd
from bsamreader import ThroughFlow
from django.conf import settings
from django.template import Context, Template
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from apps.main.models import Cas, Row, RowPair
from apps.main.modules.graph_perfo_0d.utils import (
    define_config_list,
    get_flux_alias,
    get_row_from_antarescard,
    read_input_file,
    rename_row,
)


def export_data_html(data, etat):
    # Générer le nom de fichier avec la date actuelle
    current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    new_filename = f"Perfos0D_{etat.projet.name}_{etat.name}_{current_date}.html"
    html_filepath = os.path.join(etat.work_directory, "Perfo0D", new_filename)

    template_path = os.path.join(
        settings.BASE_DIR, "apps", "templates", "trunks", "base_graph_export.html"
    )
    # Charger et lire le template HTML
    with open(template_path, encoding="utf-8") as file:
        template = Template(file.read())

    # Rendre le contenu avec les données
    data["etat"] = etat.name
    data["projet"] = etat.projet
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data["date"] = current_date
    data["created_by"] = etat.projet.created_by
    context = Context(data)
    rendered_content = template.render(context)

    # Sauvegarder le contenu HTML dans un fichier externe
    with open(html_filepath, "w", encoding="utf-8") as output_file:
        output_file.write(rendered_content)

    os.chmod(html_filepath, 0o775)


def export_data_excel(data, etat):
    data_all = {}

    # Rassembler les DataFrames par étage
    for etage, etage_data in data.items():
        frames = [
            df for df in etage_data.values()
            if df is not None and not df.empty and not df.isna().all().all()
        ]
        data_all[etage] = pd.concat(frames,
                                    ignore_index=True) if frames else pd.DataFrame()

    # Préparer le nom du fichier
    current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    new_filename = f"Perfos0D_{etat.projet.name}_{etat.name}_{current_date}.xlsx"
    excel_filepath = os.path.join(etat.work_directory, "Perfo0D", new_filename)

    # Colonnes à masquer
    column_to_mask = ['element_color', 'id', 'marker', 'is_only_stator']
    colors_by_sheet = {}

    # Export initial vers Excel
    with pd.ExcelWriter(excel_filepath) as writer:
        for sheet, df in data_all.items():
            if df.empty:
                df.to_excel(writer, sheet_name=sheet, index=False)
                continue

            # Récupérer les couleurs
            colors = df[
                'element_color'].tolist() if 'element_color' in df.columns else [
                                                                                    '#000000'] * len(
                df)
            colors_by_sheet[sheet] = colors

            # Supprimer colonnes à masquer
            cols_to_drop = [col for col in column_to_mask if col in df.columns]
            df = df.drop(columns=cols_to_drop)

            # Réordonner colonnes
            cols = df.columns.tolist()
            if 'name' in cols:
                cols.insert(0, cols.pop(cols.index('name')))
            if 'data_type' in cols:
                cols.insert(1, cols.pop(cols.index('data_type')))
            df = df[cols]

            # Tri par Qcorr **à l'intérieur des groupes**
            group_cols = ['group']  # définir les colonnes de regroupement
            if 'Qcorr' in df.columns:
                if all(col in df.columns for col in group_cols):
                    df = df.groupby(group_cols, sort=False,
                                    group_keys=False).apply(
                        lambda x: x.sort_values(by='Qcorr', ascending=True)
                    ).reset_index(drop=True)
                else:
                    df = df.sort_values(by='Qcorr', ascending=True).reset_index(
                        drop=True)

            df.to_excel(writer, sheet_name=sheet, index=False)

    # Mise en forme Excel (bordures et couleurs)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    wb = load_workbook(excel_filepath)
    for sheet, colors in colors_by_sheet.items():
        ws = wb[sheet]

        # Couleur d'en-tête
        fill_head = PatternFill(start_color='D3D3D3', end_color='D3D3D3',
                                fill_type='solid')
        for cell in ws[1]:
            cell.fill = fill_head
            cell.border = border

        # Couleur des textes par ligne
        for row_num, couleur in enumerate(colors, start=2):
            font = Font(color=couleur.replace("#", "")) if isinstance(couleur,
                                                                      str) and couleur.startswith(
                "#") else None
            for cell in ws[row_num]:
                if font:
                    cell.font = font
                cell.border = border

    wb.save(excel_filepath)
    os.chmod(excel_filepath, 0o775)


def set_data_row(cas):

    antarescard_filepath = os.path.join(cas.repertory, "calculation/carte_antares.py")

    config_dico = read_input_file(antarescard_filepath)

    row_data = {}

    for grid_key, grid_val in config_dico['gridFields'].items():
        row_data[grid_key] = {
            "flux": get_flux_alias(grid_val['Flux'], data_type='cfd'),
            "position": grid_val['FluxIndex'],
            "omega": grid_val['Omega'],
            "nb_blade": grid_val['TotalBladeCount'],
            "bsam_name": grid_val['GridBSAMName']
        }

    cas.row_metadata = row_data
    cas.save()

def set_inlet_outlet_row_obj(case_obj, rowpair_obj):

    pair_row_name_split = rowpair_obj.name.split("-")

    if len(pair_row_name_split) == 1:
        row_amont, row_aval = pair_row_name_split[0], pair_row_name_split[0]
    else:
        row_amont, row_aval = pair_row_name_split[0], pair_row_name_split[1]

    rowpair_obj.entry_row = Row.objects.get(cas=case_obj, name=row_amont)
    rowpair_obj.exit_row = Row.objects.get(cas=case_obj, name=row_aval)
    rowpair_obj.save()


def extract_average_values(hdfMoy, r_gaz, cp) -> Dict[str, float]:
    extracted = {}
    for var in ["Pta", "Tta", "Q", "Ps", "Ptr"]:
        if var not in hdfMoy:
            rho = hdfMoy['Density'][()][0]
            if "V" not in hdfMoy:
                vx = hdfMoy['MomentumX'][()][0] / rho
                vy = hdfMoy['MomentumY'][()][0] / rho
                vz = hdfMoy['MomentumZ'][()][0] / rho
                v = np.sqrt(vx ** 2 + vy ** 2 + vz ** 2)
            else:
                v = hdfMoy['V'][()][0]

            Et = hdfMoy['EnergyStagnationDensity'][()][0] / rho
            ea = Et - 0.5 * v ** 2
            ts = ea / (cp - r_gaz)
            ps = rho * r_gaz * ts
            hs = ts * cp
            ht = hs + 0.5 * v ** 2
            tt = ht / cp
            match var:
                case "Ps":
                    extracted[var] = ps
                case "Tta":
                    extracted[var] = tt
        else:
            extracted[var] = hdfMoy[var][()][0]

    return extracted

def process_grid(cas, grid, grid_name: str, plan_amont, plan_aval):
    results = []

    grid_name_bsam = str(grid["BSAM_NAME"][()].decode())
    grid_omega = grid["Omega"][()]
    r_gaz = grid["Rgaz"][()]
    cp = grid["polynomeCoeff"][()][0]
    gamma = cp / (cp - r_gaz)
    if "BSAM_Cuts" not in grid:
        return results

    for plane_name in [plan_amont, plan_aval]:

        plane_hdf = grid["BSAM_Cuts"][plane_name]

        if grid_name and grid_name.startswith('2.'):
            grid_flow = "Secondaire"
        elif grid_name and grid_name.startswith('3.'):
            grid_flow = "Primaire"
        else:
            grid_flow = "Total"

        flow = None
        if 'BSAM_' in plane_name:
            n_flow = plane_name.split('BSAM_')[-1].split('_')[0]
            flow = {"1": "Total", "2": "Secondaire", "3": "Primaire"}.get(n_flow)
        elif '_Prim' in plane_name:
            flow = "Primaire"
        elif '_Sec' in plane_name:
            flow = "Secondaire"
        elif plane_name.startswith('1.'):
            flow = "Total"
        elif plane_name.startswith('2.'):
            flow = "Secondaire"
        elif plane_name.startswith('3.'):
            flow = "Primaire"

        grid_flow_name = flow if flow else grid_flow

        for instant in plane_hdf:
            plane_instant_hdf = plane_hdf[instant]
            if "Average" not in plane_instant_hdf:
                continue

            plane_instant_0d_hdf = plane_instant_hdf["Average"]

            for moy_0d in ["Moyenne_type_5_10", "Moyenne_type_5"]:
                if moy_0d in plane_instant_0d_hdf:
                    hdfMoy = plane_instant_0d_hdf[moy_0d]
                    extracted = extract_average_values(hdfMoy, r_gaz, cp)
                    for var, value in extracted.items():
                        results.append({
                            "BSAM_Name": grid_name_bsam,
                            "Omega": grid_omega,
                            "Plane_Name": plane_name,
                            "Flow": grid_flow_name,
                            "Variable": var,
                            "Value": value,
                            "gamma": gamma,
                            "r_gaz": r_gaz,
                            "cp": cp,
                        })
                    break

    return results

def check_planes_availibility(planes, user_planes, default_planes=['Inlet', 'Outlet']):
    plane_inlet = default_planes[0]
    plane_outlet = default_planes[-1]
    for plane_name in planes:
        if user_planes[0].lower() in plane_name.lower():
            plane_inlet = plane_name
        if user_planes[-1].lower() in plane_name.lower():
            plane_outlet = plane_name

    return [plane_inlet, plane_outlet]


def load_hdf_data(cas, rowpair_name, plane_dict):
    plan_amont = cas.iso_vitesse.etat.plan_amont_selected
    plan_aval = cas.iso_vitesse.etat.plan_aval_selected
    data = []

    with h5py.File(cas.file_path, 'r') as f:
        for grid_name in f:
            if grid_name in ["CGNSLibraryVersion", "DataNozzle"]:
                continue
            grid = f[grid_name]
            planes = check_planes_availibility(grid["BSAM_Cuts"], [plan_amont, plan_aval])
            results = process_grid(cas, grid, grid_name, planes[0], planes[1])

            for element in results:
                if element["BSAM_Name"] == rowpair_name:
                    plane_dict[rowpair_name][cas.id] = planes

            data.extend(results)

    df = pd.DataFrame(data)

    row_list = [
        rename_row(x)
        for x in df["BSAM_Name"].dropna().unique().tolist()
    ]
    df["BSAM_Name"] = df["BSAM_Name"].apply(rename_row)

    return df, row_list, plane_dict


def get_hdf_moy(cas, plane):
    with h5py.File(cas.file_path, 'r') as f:
        for grid_name in f:
            if grid_name in ["CGNSLibraryVersion", "DataNozzle"]:
                continue
            grid = f[grid_name]

            for plane_name in plane:
                if "inlet" in plane_name.lower() or "outlet" in plane_name.lower():
                    plane_hdf = grid["BSAM_Cuts"][plane_name]
                    for instant in plane_hdf:
                        plane_instant_hdf = plane_hdf[instant]
                        plane_instant_0d_hdf = plane_instant_hdf["Average"]
                        for moy_0d in ["Moyenne_type_5_10", "Moyenne_type_5"]:
                            if moy_0d in plane_instant_0d_hdf:
                                extract_and_save_moy(cas, moy_0d)
                                break


def extract_and_save_moy(cas, str_moy):
    parts = str_moy.split("_")

    # Cas HDF : Moyenne_type_X ou Moyenne_type_X_Y
    if str_moy.startswith("Moyenne_type_"):
        cas.moyenne_type = str_moy.replace("Moyenne_type_", "")
        cas.save()

    # Cas Excel : Perfos0D_moy_X
    elif "_moy_" in str_moy:
        cas.moyenne_type = "5_" + parts[-1]
        cas.save()


def get_planes_on_communs(etat, lst_cas_cannelle_active_hdf):

    liste_sets_amont = []
    liste_sets_aval = []

    for cas_hdf in lst_cas_cannelle_active_hdf:
        amont_planes_cas = []
        aval_planes_cas = []

        with h5py.File(cas_hdf.file_path, "r") as f:
            for grid_name in f:
                if grid_name in ("CGNSLibraryVersion", "DataNozzle"):
                    continue

                grid = f[grid_name]
                if "BSAM_Cuts" not in grid:
                    continue

                for plane_str in grid["BSAM_Cuts"]:

                    if ("BA" in plane_str) or ("Inlet" in plane_str):
                        amont_planes_cas.append(plane_str)

                    if ("BF" in plane_str) or ("Outlet" in plane_str):
                        aval_planes_cas.append(plane_str)

        if amont_planes_cas:
            liste_sets_amont.extend(amont_planes_cas)

        if aval_planes_cas:
            liste_sets_aval.extend(aval_planes_cas)

    return liste_sets_amont, liste_sets_aval


def sort_planes(planes):

    lst_planes_sorted = []

    for plan in planes:

        if plan == "Inlet":
            lst_planes_sorted.append(((0, 0), plan))

        elif plan == "Outlet":
            lst_planes_sorted.append(((1, 0), plan))

        elif plan == "BA":
            lst_planes_sorted.append(((2, 0), plan))

        elif plan.startswith("BA-"):
            number = int(plan.split("-")[1])
            lst_planes_sorted.append(((3, number), plan))

        elif plan == "BF":
            lst_planes_sorted.append(((4, 0), plan))

        elif plan.startswith("BF+"):
            number = int(plan.split("+")[1])
            lst_planes_sorted.append(((5, number), plan))

    lst_planes_sorted.sort()

    return [plan[1] for plan in lst_planes_sorted]


def get_plane_configuration(etat):

    lst_cas_cannelle_active_hdf = Cas.objects.filter(
        iso_vitesse__etat=etat,
        iso_vitesse__file_type="hdf",
        used=True
    )

    plans_amont_communs, plans_aval_communs = (
        get_planes_on_communs(etat, lst_cas_cannelle_active_hdf))

    plane_BA = sort_planes(extract_BA_BF_planes(plans_amont_communs, "BA"))
    plane_BF = sort_planes(extract_BA_BF_planes(plans_aval_communs, "BF"))

    if etat.plan_amont != plane_BA or etat.plan_aval != plane_BF:
        etat.plan_amont = plane_BA
        etat.plan_aval = plane_BF
        for cas_hdf in lst_cas_cannelle_active_hdf:
            cas_hdf.calculate_perfo = False
            cas_hdf.save()
        etat.save()

    return plans_amont_communs, plans_aval_communs

def extract_BA_BF_planes(plans, prefix):
    result = []

    for p in plans:
        if p in ["Inlet", "Outlet"]:
            if p not in result:
                result.append(p)
            continue
        if "(" not in p:
            continue
        level = p.split("(")[1].replace(")", "")

        if level.startswith(prefix) and level not in result:
            result.append(level)

    result.sort()
    return result

def load_bdd_data(cas, pair_row):
    rp = RowPair.objects.get(cas=cas, name=pair_row)

    data = {
        "Qcorr_ref": rp.Qcorr_ref,
        "Pi": rp.Pi,
        "Qcorr": rp.Qcorr,
        "Tau": rp.Tau,
        "Cd": rp.Cd,
        "Etapol": rp.Etapol,
        "KD": cas.iso_vitesse.recalage_kd,
        "PisQcorr_ref": (
            rp.Pi / rp.Qcorr_ref
            if (rp.Pi is not None and rp.Qcorr_ref not in (0, None))
            else None
        )
    }

    return data, rp

def get_value_from_df_hdf(df, name, plane, variable):

    values = df[
        (df['BSAM_Name'] == name) &
        (df['Plane_Name'] == plane) &
        (df['Variable'] == variable)
    ]['Value'].values

    return values[0] if len(values) > 0 else None


def set_type_of_rowpair(rowpair, sheet_name):
    if sheet_name == 'Global':
        rowpair.type = 'global'
    elif re.search(r'RM\d+-RD\d+', rowpair.name):
        rowpair.type = 'etage'
    elif re.search(r'RD\d+-RM\d+', rowpair.name):
        rowpair.type = 'pseudo_etage'
    else:
        rowpair.type = 'isole'
    rowpair.save()


def get_row_from_bsam(bsam):

    tf = ThroughFlow.from_bsam(bsam.file_path)

    for zone in tf.zones:
        for index, row in enumerate(zone.rows):
            row_obj, _ = Row.objects.get_or_create(
                flux=get_flux_alias(row.inlet.jflux, data_type='bsam'),
                position=index + 1,
                omega=row.omega,
                nb_blade=row.z,
                bsam_name=row.name,
                name=rename_row(row.name),
                cas=bsam,
            )

            row_obj.type = 'rotor' if row_obj.omega == 0.0 else 'stator'
            row_obj.save()

    return Row.objects.filter(cas=bsam)

def get_row_from_antares_card(case_obj):

    antarescard_filepath = os.path.join(case_obj.repertory, "calculation/carte_antares.py")

    config_dico = read_input_file(antarescard_filepath)

    for grid_key, grid_val in config_dico['gridFields'].items():
        row_obj, _ = Row.objects.get_or_create(
            flux=get_flux_alias(grid_val['Flux'], data_type='cfd'),
            position=grid_val['FluxIndex'],
            omega=grid_val['Omega'],
            nb_blade=grid_val['TotalBladeCount'],
            bsam_name=grid_val['GridBSAMName'],
            name=rename_row(grid_val['GridBSAMName']),
            cas=case_obj,
        )

        row_obj.type = 'rotor' if row_obj.omega == 0.0 else 'stator'
        row_obj.save()

    return Row.objects.filter(cas=case_obj)

def get_iso_config(iso_vitesse):

    if iso_vitesse.get_lst_row_config():
        return iso_vitesse.get_lst_row_config()

    else:
        for case in Cas.objects.filter(iso_vitesse=iso_vitesse):

            antares_file = os.path.join(
                case.repertory,
                "calculation/carte_antares.py",
            )

            config_dico = read_input_file(antares_file)
            row_data = {}
            for grid_key, grid_val in config_dico['gridFields'].items():
                row_data[grid_key] = {
                    "flux": get_flux_alias(grid_val['Flux'], data_type='cfd'),
                    "position": grid_val['FluxIndex'],
                    "omega": grid_val['Omega'],
                    "nb_blade": grid_val['TotalBladeCount'],
                    "bsam_name": grid_val['GridBSAMName']
                }

            case.row_metadata = row_data
            case.save()

        row_list_anna = get_row_from_antarescard(antares_file)
        iso_vitesse.row_config = define_config_list(row_list_anna, rename=True)
        iso_vitesse.save(update_fields=["row_config"])
        return iso_vitesse.get_lst_row_config()

def make_json_serializable(obj):
    if isinstance(obj, (np.integer, np.floating)):
        return obj.item()
    elif isinstance(obj, (datetime,)):
        return obj.isoformat()
    elif hasattr(obj, '__dict__'):
        return str(obj)  # fallback
    return obj


def get_sheet_name_from_rowpair(file_path, rowpair_obj):

    df = pd.read_excel(file_path, sheet_name=None)

    pair_row_name_split = rowpair_obj.name.split("-")

    if len(pair_row_name_split) == 1:
        row_amont, row_aval = pair_row_name_split[0], pair_row_name_split[0]
    else:
        row_amont, row_aval = pair_row_name_split[0], pair_row_name_split[1]

    for sheet in df.keys():
        upstream_plane = df[sheet].get('plan_amont')
        downstream_plane = df[sheet].get('plan_aval')
        row_name_up, row_name_down = "", ""
        if upstream_plane is not None:
            row_name_up = rename_row(upstream_plane.iloc[0].split('_')[0])
        if downstream_plane is not None:
            row_name_down = rename_row(downstream_plane.iloc[0].split('_')[0])

        if row_name_up == row_amont and row_name_down == row_aval:
            return sheet

def make_planes_tab_data(all_plane_dict):
    # récupérer tous les ids existants
    all_cas = set()
    for rowpair in all_plane_dict:
        for cas in all_plane_dict[rowpair]:
            all_cas.add(cas)

    # compléter les ids manquants avec valeur par défaut
    for rowpair in all_plane_dict:
        for cas in all_cas:
            try:
                if not all_plane_dict[rowpair][cas]:
                    all_plane_dict[rowpair][cas] = ['Inlet', 'Outlet']
            except:
                pass
    return all_plane_dict